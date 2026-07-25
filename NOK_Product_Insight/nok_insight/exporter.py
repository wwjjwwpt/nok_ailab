from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .metrics import AnalysisResult, METRIC_DEFINITIONS


COLORS = {
    "navy": "14283D",
    "teal": "0F766E",
    "amber": "D97706",
    "red": "C94C4C",
    "paper": "F4F7F5",
    "light_teal": "DDF3EF",
    "light_amber": "FEF1D6",
    "light_red": "FBE3E3",
    "line": "D8E0E5",
    "white": "FFFFFF",
}

FIELD_NAMES = {
    "rank": "销售排名",
    "product_id": "品号",
    "product_name": "品名",
    "product_type": "产品类型",
    "spec": "规格",
    "material": "材料",
    "note": "备注",
    "reference_price": "参考售价",
    "monthly_sales_amount": "月均销售金额",
    "total_sales_qty": "总销售数量",
    "avg_monthly_qty": "月均销售数量",
    "total_inventory": "总在库数量",
    "total_available": "总可用量",
    "unfinished_qty": "未完成数量",
    "purchased_unrecorded": "已采购未录入",
    "unfinished_total": "合计未完成",
    "inventory_cover": "库存可销售月",
    "available_cover": "可用量可销售月",
    "unfinished_cover": "未完成可销售月",
    "total_cover": "合计可销月数",
    "target_months": "目标可销月数",
    "minimum_order_qty": "最低补货量",
    "sales_share": "销售占比",
    "inventory_share": "库存占比",
    "last_inbound_date": "最后进货时间",
    "last_inbound_qty": "最后进货数量",
    "customer_count": "购买客户数",
    "purchase_count": "购买次数",
    "monthly_purchase_qty": "当月采购数量",
    "monthly_purchase_amount": "当月采购金额",
    "recent_purchase_date": "最近采购时间",
    "recent_purchase_qty": "最近采购数量",
    "stockout_count": "断货次数",
}


def _safe_value(value: object) -> object:
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    return value


def _style_header(worksheet, row: int, start_col: int, end_col: int) -> None:
    fill = PatternFill("solid", fgColor=COLORS["navy"])
    font = Font(color=COLORS["white"], bold=True)
    for cell in worksheet.iter_cols(
        min_col=start_col, max_col=end_col, min_row=row, max_row=row
    ):
        target = cell[0]
        target.fill = fill
        target.font = font
        target.alignment = Alignment(horizontal="center", vertical="center")


def _autowidth(worksheet, min_width: int = 10, max_width: int = 38) -> None:
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        width = min_width
        for cell in column_cells[:200]:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        worksheet.column_dimensions[letter].width = width


def export_analysis(result: AnalysisResult, destination: str | Path) -> Path:
    output = Path(destination).expanduser().resolve()
    output.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result.products.to_excel(writer, sheet_name="产品指标", index=False)
        result.alerts.to_excel(writer, sheet_name="风险预警", index=False)
        result.source.data.to_excel(writer, sheet_name="原始数据", index=False)
        mappings = [
            {
                "标准字段": FIELD_NAMES.get(field, field),
                "原Excel字段": column,
                "识别类型": "业务字段",
            }
            for field, column in result.source.columns.fields.items()
        ]
        mappings.extend(
            {
                "标准字段": month.strftime("%Y-%m销售数量"),
                "原Excel字段": column,
                "识别类型": "月度销量",
            }
            for month, column in result.source.columns.monthly_sales
        )
        pd.DataFrame(mappings).to_excel(writer, sheet_name="字段映射", index=False)
        pd.DataFrame(METRIC_DEFINITIONS, columns=["指标", "计算方法", "业务用途"]).to_excel(
            writer, sheet_name="指标说明", index=False
        )
        pd.DataFrame({"数据质量提示": result.quality_messages or ["未发现明显问题"]}).to_excel(
            writer, sheet_name="数据质量", index=False
        )

        workbook = writer.book
        summary_sheet = workbook.create_sheet("管理摘要", 0)
        summary_sheet.sheet_view.showGridLines = False
        summary_sheet["A1"] = "产品经营分析摘要"
        summary_sheet["A1"].font = Font(size=20, bold=True, color=COLORS["navy"])
        summary_sheet["A2"] = f"来源：{result.source.path.name}｜工作表：{result.source.sheet_name}"
        summary_sheet["A2"].font = Font(size=10, color="607080")
        summary_sheet.merge_cells("A1:F1")
        summary_sheet.merge_cells("A2:F2")

        summary_items = list(result.summary.items())
        for idx, (label, value) in enumerate(summary_items, start=4):
            summary_sheet.cell(idx, 1, label)
            summary_sheet.cell(idx, 2, _safe_value(value))
            summary_sheet.cell(idx, 1).font = Font(bold=True, color=COLORS["navy"])
            summary_sheet.cell(idx, 2).font = Font(size=12, bold=True, color=COLORS["teal"])
            summary_sheet.cell(idx, 1).fill = PatternFill("solid", fgColor=COLORS["paper"])
            summary_sheet.cell(idx, 2).fill = PatternFill("solid", fgColor=COLORS["paper"])

        category_counts = (
            result.products["产品分层"].value_counts().rename_axis("产品分层").reset_index(name="产品数")
        )
        start_row = 4
        summary_sheet.cell(start_row, 4, "产品分层")
        summary_sheet.cell(start_row, 5, "产品数")
        for offset, row in enumerate(category_counts.itertuples(index=False), start=1):
            summary_sheet.cell(start_row + offset, 4, row[0])
            summary_sheet.cell(start_row + offset, 5, int(row[1]))
        _style_header(summary_sheet, start_row, 4, 5)

        top_columns = [
            "品号",
            "品名",
            "产品分层",
            "经营优先分",
            "月均销售金额(万元)",
            "合计可销月数",
            "近3个月增长率",
            "建议动作",
        ]
        top = result.products.nlargest(10, "经营优先分")[top_columns]
        table_row = max(16, len(summary_items) + 6)
        summary_sheet.cell(table_row, 1, "优先处理的10个产品")
        summary_sheet.cell(table_row, 1).font = Font(size=14, bold=True, color=COLORS["navy"])
        for col_index, column in enumerate(top_columns, start=1):
            summary_sheet.cell(table_row + 1, col_index, column)
        _style_header(summary_sheet, table_row + 1, 1, len(top_columns))
        for row_index, values in enumerate(top.itertuples(index=False), start=table_row + 2):
            for col_index, value in enumerate(values, start=1):
                summary_sheet.cell(row_index, col_index, _safe_value(value))

        for sheet_name in ("产品指标", "风险预警", "原始数据", "字段映射", "指标说明", "数据质量"):
            sheet = workbook[sheet_name]
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            sheet.sheet_view.showGridLines = False
            _style_header(sheet, 1, 1, sheet.max_column)
            sheet.row_dimensions[1].height = 28
            _autowidth(sheet)

        product_sheet = workbook["产品指标"]
        header_lookup = {cell.value: cell.column for cell in product_sheet[1]}
        for name in ("近3个月增长率", "销售贡献率", "断货率"):
            if name in header_lookup:
                letter = get_column_letter(header_lookup[name])
                for cell in product_sheet[letter][1:]:
                    cell.number_format = "0.0%"
        for name in ("合计可销月数", "目标可销月数", "库存覆盖缺口", "需求波动系数"):
            if name in header_lookup:
                letter = get_column_letter(header_lookup[name])
                for cell in product_sheet[letter][1:]:
                    cell.number_format = "0.0"
        if "经营优先分" in header_lookup:
            col = get_column_letter(header_lookup["经营优先分"])
            product_sheet.conditional_formatting.add(
                f"{col}2:{col}{product_sheet.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color=COLORS["light_teal"],
                    mid_type="percentile",
                    mid_value=60,
                    mid_color=COLORS["light_amber"],
                    end_type="max",
                    end_color=COLORS["light_red"],
                ),
            )

        alert_sheet = workbook["风险预警"]
        for row in range(2, alert_sheet.max_row + 1):
            level = alert_sheet.cell(row, 1).value
            color = {
                "高": COLORS["light_red"],
                "中": COLORS["light_amber"],
                "低": COLORS["light_teal"],
            }.get(level, COLORS["paper"])
            for cell in alert_sheet[row]:
                cell.fill = PatternFill("solid", fgColor=color)

        thin = Side(style="thin", color=COLORS["line"])
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="center", wrap_text=False)
            if sheet.title == "管理摘要":
                for row in sheet.iter_rows(min_row=4, max_row=min(13, sheet.max_row), min_col=1, max_col=2):
                    for cell in row:
                        cell.border = Border(bottom=thin)
            _autowidth(sheet)

    return output
