from __future__ import annotations

import unittest
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from nok_insight.exporter import export_analysis
from nok_insight.loader import load_excel
from nok_insight.metrics import analyze

from .helpers import create_sample_workbook


class ExporterTests(unittest.TestCase):
    def test_exports_management_workbook(self) -> None:
        with TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = create_sample_workbook(root / "sample.xlsx")
            result = analyze(load_excel(source))
            output = export_analysis(result, root / "analysis.xlsx")
            self.assertTrue(output.exists())
            workbook = load_workbook(output, read_only=True, data_only=False)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["管理摘要", "产品指标", "风险预警", "原始数据", "字段映射", "指标说明", "数据质量"],
                )
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
